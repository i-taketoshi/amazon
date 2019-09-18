# -*- coding: utf-8 -*-
"""
Created on Sun Feb  4 15:51:49 2018

@author: taketoshi

保存されているUSのwebページを解析し、
購入した商品情報をエクセルへインプットする

webページの保存は手動で実施する

"""
import sys
import time
from bs4 import BeautifulSoup
import os
import openpyxl as px
import re
from copy import copy

path = r'D:\workspace\利益管理\Order\\'
files = []
filelists = [] 
for x in os.listdir(path):
    if os.path.isfile(path + x):  #isdirの代わりにisfileを使う
        files.append(x)
        
#実行済み一覧を保持しているテキストを開く    
f = open('D:\workspace\利益管理\Order\OrderList.txt', 'r')
#一覧情報を配列へ格納
for line in f:
    filelists.append(line.replace('\n',''))
f.close()

#重複削除
set_ab = set(files) - set(filelists)
print(set_ab)

for s in set_ab:
    
    Filename = r'D:\workspace\利益管理\research_table.xlsx'
    # ブックを開く
    wb = px.load_workbook(Filename, read_only=False)
    #シートを固定
    ws = wb["購入"]
    #最終行取得
    max_row = 1
    while ws.cell(row = max_row, column = 1).value != None:
        max_row += 1
    
    
    
    addname = s
    # ファイルをオープンする
    text_data = open(r"D:\workspace\利益管理\Order\\" + addname, "rb")
    # すべての内容を読み込む
    contents = text_data.read()
    # ファイルをクローズする
    text_data.close()
    
    soup = BeautifulSoup(contents,"lxml")
    
    
    #購入日やオーダー番号等のインヴォイス情報取得
    invoice = soup.find_all(class_="order-date-invoice-item")
    invarr = []
    for inv in invoice:
        tmp = inv.text.strip().replace("Ordered on ",'').strip()
        tmp = tmp.replace("Order#",'').strip()
        invarr.append(tmp)
    
    
    
    pattern = r"/\w{10}/"
    repatter = re.compile(pattern)
    grid_inner = soup.find_all(class_="a-fixed-left-grid-inner")
    print(invarr)
    for b in grid_inner:
        #US　ASIN入力
        atag = b.find("a",class_="a-link-normal")
        matchOB = repatter.search(atag.get("href"))
        asin = matchOB.group(0).replace('/','')
        #ASIN
        ws.cell(row = max_row, column = 3).value = asin
        
        #No入力
        ws.cell(row = max_row, column = 1).value = max_row - 1
    
        #US価格入力
        ws.cell(row = max_row, column = 8).value =  b.find(class_='a-size-small a-color-price').string.strip().replace('$','')
        
        #個数入力
        try:
            ws.cell(row = max_row, column = 18).value = b.find(class_='item-view-qty').text.strip()
        except:
            ws.cell(row = max_row, column = 18).value = 1
        #購入日           
        ws.cell(row = max_row, column = 19).value = invarr[0]
        #オーダー番号
        ws.cell(row = max_row, column = 27).value = invarr[1]
               
        max_row += 1
        
        #書式のコピー
        for c in range(ws.max_column):   
            ws.cell(row = max_row - 1, column = c+1)._style = copy(ws.cell(row = 2, column = c+1)._style)
            #ws.cell(row = max_row_exp, column = c+1)._style = copy(ws.cell(row = 2, column = c+1)._style)
        
    
    wb.save(Filename)
    
    ws = wb["経費"]
    #最終行取得
    max_row_exp = 1
    while ws.cell(row = max_row_exp, column = 1).value != None:
        max_row_exp += 1
    
    ws.cell(row = max_row_exp, column = 1).value = max_row_exp - 1#No
    ws.cell(row = max_row_exp, column = 2).value = "US amazonでの商品仕入"
    ws.cell(row = max_row_exp, column = 3).value = "=D" + str(max_row_exp) + "*データ!$A$2"
    ws.cell(row = max_row_exp, column = 5).value = invarr[0]#購入日
    ws.cell(row = max_row_exp, column = 6).value = invarr[1]#オーダー番号
    
    subtotals = soup.find_all(class_="a-color-base a-text-bold")
    for s in subtotals:
        if s.string.find('$') > 0:
            ws.cell(row = max_row_exp, column = 4).value = s.string.replace('$','').strip()#合計金額
    
            break
    
    #書式のコピー
    #for c in range(ws.max_column):    
    #    ws.cell(row = max_row_exp, column = c+1)._style = copy(ws.cell(row = 2, column = c+1)._style)
    
    wb.save(Filename)
    
#一覧情報を上書きして保存    
f = open('D:\workspace\利益管理\Order\OrderList.txt', 'w')  #書き込みモードでオープン 
f.writelines("\n".join(files))
f.close()

print("end")
    
