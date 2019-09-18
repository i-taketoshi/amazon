# -*- coding: utf-8 -*-
"""
Created on Sun Feb  4 15:51:49 2018

@author: taketoshi

OrderInputを実行した後に実行する

エクセルに入力されているASINを元に詳細商品情報を入力する


"""
import sys
import time
from bs4 import BeautifulSoup

import openpyxl as px
from pprint import pprint
from copy import copy
import mwstestClass

def get_price_jp(p):
    if p.find("landedprice") is None:
        return "None"
    else:
        return p.find("landedprice").find("amount").string
def isNone(p,whl):
    try:
        
        if p.find("ns2:packagedimensions").find(whl) is None:
            if p.find("ns2:itemdimensions").find(whl) is None:
                return 0
            else:
                return p.find("ns2:itemdimensions").find(whl).string
            
        else:
            return p.find("ns2:packagedimensions").find(whl).string
    except:
        return 0
            
        

def chunked(iterable, n):
    return [iterable[x:x + n] for x in range(0, len(iterable), n)]


    
start = time.time()


SellerID           = "A2ILL9FJMGGGDJ"
AWSAccesKeyId      = "AKIAIA4ALUSZTLIJU2IQ"
AWSSecretAccessKey = "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"

product = mwstestClass.Products(
    AWSAccessKeyId=AWSAccesKeyId,
    AWSSecretAccessKey=AWSSecretAccessKey,
    SellerId=SellerID,
    Region='JP')
MarketplaceId      = product.domain#A1VC38T7YXB528
  
Filename = r'D:\workspace\利益管理\research_table調査用.xlsx'


# ブックを開く
wb = px.load_workbook(Filename, read_only=False)
#シートを固定
ws = wb["購入"]
#最終行取得
max_row = 1

asins = []
excel = []

#ASINとセル位置を取得
#US ASINがあるものを検索
while ws.cell(row = max_row, column = 3).value != None:
    #日本価格が空白のものを選択
    if ws.cell(row = max_row, column = 6).value is None:
        #アメリカASIN
        asin = ws.cell(row = max_row, column = 3).value
        #書き込み用配列にストック
        excel.append({'ASIN':asin, 'cellno':max_row})
        asins.append(asin)
    max_row += 1


#ASINを5分割.7分割以上にするとMWSが失敗する
splitnum = 5
asin10 = chunked(asins, splitnum)


#MWSを5個単位のASINで実行
for num, value in enumerate(asin10):    
    #商品情報を取得
    print(value)
    prices = []
    #カート価格を取得
    response = product.get_competitive_pricing_for_asin(MarketplaceId, value)
    soup = BeautifulSoup(response,"lxml")
    pricings = soup.findAll("getcompetitivepricingforasinresult")
    #print(soup.prettify())
    for pnum, p in enumerate(pricings):
        
        #書き込み用配列にストック
        if p.find("amount") is not None:
            excel[num*splitnum + pnum].update({"price":p.find("amount").string})
            prices.append(p.find("amount").string)
        else:
            excel[num*splitnum + pnum].update({"price":""})
            prices.append(0)

    #FBA手数料を取得
   
    response = product.get_myfees_estimate(MarketplaceId, value, prices)
    soup = BeautifulSoup(response,"lxml")

    
    for fnum, fee in enumerate(soup.findAll('feesestimateresult')):
        
        #書き込み用配列にストック
        if fee.find("totalfeesestimate") is not None:
            excel[num*splitnum + fnum].update({"fee":fee.find("totalfeesestimate").find('amount').string})
        else:
            excel[num*splitnum + fnum].update({"fee":""})


    response = product.get_matching_product_forid(MarketplaceId, value)
    soup = BeautifulSoup(response,"lxml")  

    #print(soup.prettify())
    for pdtnum, p in enumerate(soup.findAll("getmatchingproductforidresult")):

        if p.find("ns2:binding") is not None:
            category = p.find("ns2:binding").string
        else:
            category = ''
        if p.find("rank"):
            rank = p.find("rank").string
        else:
            rank = ""
            
        if p.find("ns2:title") is not None:
            title = p.find("ns2:title").string
        else:
            title = ''
        
        height = float(isNone(p,"ns2:height")) * 2.54
        length = float(isNone(p,"ns2:length")) * 2.54
        width = float(isNone(p,"ns2:width")) * 2.54
        weight = float(isNone(p,"ns2:weight")) * 453.592
        postage = 2020 * ((height + length + width)/160) * 0.8
        
        '''
        if p.find("ns2:packagedimensions") is not None:
            
            height = float(isNone(p,"ns2:height") * 2.54)
            length = float(isNone(p,"ns2:length") * 2.54)
            width = float(isNone(p,"ns2:width") * 2.54)
            weight = float(isNone(p,"ns2:weight") * 453.592)
            
            #height = float(p.find("ns2:packagedimensions").find("ns2:height").string) * 2.54#センチ
            #length  = float(p.find("ns2:packagedimensions").find("ns2:length").string) * 2.54#センチ
            #width  = float(p.find("ns2:packagedimensions").find("ns2:width").string) * 2.54#センチ
            #weight  = float(p.find("ns2:packagedimensions").find("ns2:weight").string) * 453.592#g
            #160サイズを元にした送料目安。0.8は係数
            postage = 2020 * ((height + length + width)/160) * 0.8
        else:
            height = 0
            length  = 0
            width  = 0
            weight  = 0
            
            #160サイズを元にした送料目安。0.8は係数
            postage = 0

        '''
        
        #title = title.replace(':','')
        idx = num*splitnum + pdtnum
        excel[idx].update({"category":category})
        excel[idx].update({"title":title})
        excel[idx].update({"weight":round(weight,2)})
        excel[idx].update({"height":round(height,2)})
        excel[idx].update({"length":round(length,2)})
        excel[idx].update({"width":round(width,2)})
        excel[idx].update({"postage":round(postage,2)})
        excel[idx].update({"rank":rank})
        #print("num:{}, splitnum{}, pdtnum{}".format(num,splitnum,pdtnum))
        
        
        cellrow = excel[idx]['cellno']
        ws.cell(row = cellrow, column = 4).value = excel[idx]['title']
        ws.cell(row = cellrow, column = 5).value = excel[idx]['category']
        ws.cell(row = cellrow, column = 6).value = excel[idx]['price']
        ws.cell(row = cellrow, column = 9).value = excel[idx]['fee']
        ws.cell(row = cellrow, column = 12).value = excel[idx]['postage']
        ws.cell(row = cellrow, column = 22).value = excel[idx]['weight']
        ws.cell(row = cellrow, column = 23).value = excel[idx]['height']
        ws.cell(row = cellrow, column = 24).value = excel[idx]['length']
        ws.cell(row = cellrow, column = 25).value = excel[idx]['width']
        ws.cell(row = cellrow, column = 26).value = excel[idx]['postage']
        #ws.cell(row = cellrow, column = 32).value = excel[idx]['rank']
        
        
        time.sleep(5)
    print("{}/{}終了".format((num+1)*5, len(excel)))
    wb.save(Filename)


#for tate in range(excel[0]['cellno'],max_row):
    
    #for yoko in range(1,28):
        
        #ws.cell(row = tate, column = yoko)._style = copy(ws.cell(row = 2, column = yoko)._style)

wb.save(Filename)        

    #print("{}/{}終了".format(cnt, max_row))
#pprint(excel)
'''
#エクセルに書き込み
for e in excel:
    
    ws.cell(row = e['cellno'], column = 4).value = e['title']
    ws.cell(row = e['cellno'], column = 5).value = e['category']
    ws.cell(row = e['cellno'], column = 6).value = e['price']
    ws.cell(row = e['cellno'], column = 9).value = e['fee']
    ws.cell(row = e['cellno'], column = 12).value = e['postage']
    ws.cell(row = e['cellno'], column = 22).value = e['weight']
    ws.cell(row = e['cellno'], column = 23).value = e['height']
    ws.cell(row = e['cellno'], column = 24).value = e['length']
    ws.cell(row = e['cellno'], column = 25).value = e['width']
    ws.cell(row = e['cellno'], column = 26).value = e['postage']

for c in range(1,28):
    ws.cell(row = max_row, column = c)._style = copy(ws.cell(row = 2, column = c)._style)

wb.save(Filename)    
'''

elapsed_time = time.time() - start
print ("elapsed_time:{0}".format(elapsed_time) + "[sec]")
    
