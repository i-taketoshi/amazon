# -*- coding: utf-8 -*-
"""
Created on Wed Apr 18 00:43:48 2018

@author: taketoshi

価格差のある商品をDBから検索し、エクセルへ書き込む
書き込んだエクセルを読み込み、モノレートをスクレイピングする

"""
from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException

import time
import openpyxl as px
from random import randint
import sys
import datetime

import re
import shutil

import os
import urllib.parse

#ついでにASINも集める処理
def getASIN(localsoup):
    pass
            
def browserCheck(asin):

    #タイムアウト対策3回まで
    i = 0
    while i < 3:
        
        try:
            driver.get('https://www.amazon.co.jp/gp/product/' + asin)
            time.sleep(5)
            break
        except TimeoutException:
            i = i + 1
            print('Timeout, Retrying... (%(i)s/%(max)s)' % {'i': i, 'max': 3})
            continue
        except:
            i = 3
            print('ページの読み込み時に何かしらの失敗した')
            time.sleep(10)
            continue
    
    data = driver.page_source
    soup = BeautifulSoup(data,"lxml")
    
    star = ''
    review_num = ''
    maker = ''
    price = ''
    price_ship = ''
    availability = ''
    rank = ''
    seller_name = ''
    if soup.find(id='acrPopover') is not None:
        star = soup.find(id='acrPopover').get_text().strip()
        star = star.replace('5つ星のうち ','')
    
    if soup.find(id='acrCustomerReviewText') is not None:
        review_num = soup.find(id='acrCustomerReviewText').get_text().strip()
        review_num = review_num.replace('件のカスタマーレビュー','')

    #tittle = soup.find(id='title').get_text().strip()
    if soup.find(id='bylineInfo') is not None:    
        maker = soup.find(id='bylineInfo').get_text().strip()
    #not_prime = soup.find(id='alternativeOfferEligibilityMessaging_feature_div').get_text().strip()
    
    if soup.find(id='priceblock_ourprice') is not None:
        price = soup.find(id='priceblock_ourprice').get_text().strip()#商品価格
    
    if soup.find(id='soldByThirdParty') is not None:
        price_ship = soup.find(id='soldByThirdParty').get_text().strip()#カートに記載されている送料。存在しない場合もある。
    
    if soup.find(id='merchant-info') is not None:
        sales_ships = soup.find(id='merchant-info').findAll('a')#誰が発送する。<a>XX</a>が販売し、<a>YY</a>が発送します
    
    if soup.find(id='availability') is not None:
        availability = soup.find(id='availability').get_text().strip()#在庫切れ、予約、発送にかかるまでの期間
    
    if soup.find(id='SalesRank') is not None:
        rank = soup.find(id='SalesRank').get_text().strip()

    #print("評価：" + star)
    #print("評価数：" + review_num)
    #print("商品名：" + tittle)
    #print("メーカー名：" + maker)
    
    m_p = re.search(r"[0-9]*,*[0-9]+", price)

    if m_p:
        price = m_p.group(0).replace(',','')
        #print("商品価格：" + price)
    
    #m_p_s = re.search(r"[0-9]*,*[0-9]+", price_ship)
    m_p_s = re.search("\+ ￥ [0-9]*,*[0-9]+", price_ship)
    if m_p_s:
        price_ship = m_p_s.group(0).replace('+ ￥ ','')
        price_ship = price_ship.strip()
    elif '関東への配送料無料' in price_ship:
        price_ship = ""
    else:
        price_ship = '未定義'

    
    if '予約' in availability:
        availability = '予約商品'
    elif '在庫切れ' in availability:
        availability = '在庫切れ'
    elif '出品者' in availability:
        availability = 'カート獲得者無し'
    else:
        availability = 'カート獲得者あり'

    #print("利用状態" + availability)
    
    #販売・発送が異なる:B07D3NPTQJ
    #販売・発送が同じ：B07C92BSSN
    #予約商品：B07C3TG6FC
    #Amazon販売・発送：B009CMIHPC
    #FBA：B076DXSYDJ
    
    seller_text = sales_ships[0].get_text().strip()
    
    if 'Amazon' in seller_text:
        amazon_flag = True
        availability = 'FBA'
    else:
        amazon_flag = False
        #アルファベットのメーカーを判定
        pattern = "[a-zA-Z0-9_-]+"
        #アルファベット部分抽出
        matchOB = re.search(pattern , maker)
        #メーカー販売か、セラー販売か区別する
        if matchOB:#アルファベットメーカー
            maker_name = matchOB.group()
            if maker_name.lower() in seller_text.lower():
                #print("メーカー自己販売")
                maker_flag = True
            else:
                #print("セラー販売")
                maker_flag = False
        else:#日本語メーカー
            if maker in seller_text:
                #print("メーカー自己販売2")
                maker_flag = True
            else:
                #print("セラー販売2")
                maker_flag = False
    if len(sales_ships) == 1:
        if amazon_flag:
            seller = 'Amazon'
            shipper = seller
            availability = 'FBA'
        elif maker_flag:
            seller = 'メーカー'
            shipper = seller
        else:
            seller = 'セラー'
            shipper = seller
    else:
        seller_text2 = sales_ships[1].get_text().strip()
        if 'Amazon' in seller_text2:
            shipper = 'Amazon'
            if maker_flag:
                seller = 'メーカー'
            elif maker_flag == False:
                seller = 'セラー'
            else:
                seller = '未定義'
        else:
            shipper = '未定義'

    judge_flag = "○"
    if amazon_flag:
        judge_flag = '×'
    elif maker_flag:
        judge_flag = '×'
        
        #除外リストにメーカー名自動追記
        exclusion_path = r"D:\workspace\アマゾン出品大学\exclusion_list2.txt"
        with open(exclusion_path, "r") as f:
            exclusion_list =  f.read().split('\n')
        
        with open(exclusion_path, "a") as f:
            if maker not in exclusion_list:
                #f.write(seller + "\n")
                pass
    else:
        judge_flag = '○'

    #順位取得
    m = re.search(r"[0-9]*,*[0-9]+位", rank)
    if m:
        #print(m.group(0))
        rank = m.group(0).replace(',','')
        rank = rank.replace('位','')
        #print(rank)
        

    return seller, shipper, judge_flag, star, review_num, rank, availability, seller_name, price_ship



##########################################

#除外リスト読み込み
f = open('D:\workspace\アマゾン出品大学\exclusion_list2.txt', 'r')
exclusion_list =  f.read().split('\n')
#print(exclusion_list)
f.close()



totalstart = time.time()
path = r'D:\workspace\amazon\chromedriver.exe'
options = Options()
options.add_argument("--disable-extentions")
options.add_argument("--ignore-certificate-errors")
#options.add_argument("--headless")

###### debug ###########################
driver = webdriver.Chrome(chrome_options=options)
asin = "B01G36GDZK"
result = browserCheck(asin)
print(result)
driver.quit()
sys.exit()
###########################

todaydetail  =    datetime.datetime.today()
yymmdd = todaydetail.strftime("%Y-%m-%d")
#filename = "TH_Recommended_Item_List"
filename = "TH_Recommended_Item_List_20180615"
fol = "D:\workspace\アマゾン出品大学\\"

file_path = fol + filename + "_" + yymmdd +".xlsx"
file_path = r"D:\workspace\アマゾン出品大学\TH_Recommended_Item_List_20180615_2018-06-17.xlsx"
'''
#元のファイルは残し、コピー先のファイル名がすでにある場合は、_1とする
if os.path.isfile(file_path):
    file_path = fol + filename + "_" + yymmdd +"_1.xlsx"

shutil.copyfile(fol + filename + ".xlsx", file_path)
'''


# ブックを開く
wb = px.load_workbook(file_path, read_only=False)
#シートを固定
ws = wb["売れ筋　先週"]
#最終行取得
max_row = ws.max_row 
st = 7813

ws.cell(row = st-1, column = 15).value = "販売者"
ws.cell(row = st-1, column = 16).value = "発送者"
ws.cell(row = st-1, column = 17).value = "判定"
ws.cell(row = st-1, column = 18).value = "評価"
ws.cell(row = st-1, column = 19).value = "評価数"
ws.cell(row = st-1, column = 20).value = "順位"
ws.cell(row = st-1, column = 21).value = "状態"
ws.cell(row = st-1, column = 22).value = "販売者"
ws.cell(row = st-1, column = 23).value = "送料"

       
driver = webdriver.Chrome(chrome_options=options)
cnt = 0
for c,a in enumerate(range(st,max_row+1)):
    start = time.time()
    #return seller, shipper, judge_flag, star, review_num, rank, availability, sellername, price_ship
    
    asin = ws.cell(row = a,column = 3).value
    #amazon直販じゃないものかつ、除外リスト以外のものを実行
    if ws.cell(row = a,column = 12).value != "〇" and ws.cell(row = a,column = 6).value not in exclusion_list and \
              ws.cell(row = a,column = 15).value == None:
        
        try:
            cnt+=1
            result = browserCheck(asin)
            print('実質ブラウザチェック{}個目終了({})'.format(cnt,asin))
        
        except:
            result = ['err','err','err','err','err','err','err','err','err']
            print(asin + "エラー")
        
        ws.cell(row = a, column = 15).value = result[0]#販売者
        ws.cell(row = a, column = 16).value = result[1]#"発送者"
        ws.cell(row = a, column = 17).value = result[2]#"判定"
        ws.cell(row = a, column = 18).value = result[3]#"星"
        ws.cell(row = a, column = 19).value = result[4]#"評価数"
        ws.cell(row = a, column = 20).value = result[5]#"順位"
        ws.cell(row = a, column = 21).value = result[6]#"状態"
        ws.cell(row = a, column = 22).value = result[7]#販売者
        ws.cell(row = a, column = 23).value = result[8]#送料
        
        wb.save(file_path)
        
        elapsed_time = time.time() - start
        print ("elapsed_time:{0}".format(elapsed_time) + "[sec]")
    elif ws.cell(row = a,column = 15).value != None:
        #すでにスクレイピング済みの場合
        print("exed")
        pass
    
    else:
        ws.cell(row = a, column = 15).value = "除外リスト"#販売者
        
    print('{}個目終了({})'.format(c+1,asin))
    
    if cnt == 300:
        #break
        driver.quit()
        print('time sleep中。開始時刻{}'.format(datetime.datetime.now()))
        rest = randint(3660,5000)
        time.sleep(rest)
        print(rest)
        driver = webdriver.Chrome(chrome_options=options)
        cnt = 0
        
driver.quit()
elapsed_total_time = time.time() - totalstart
print ("総実行時間:{0}".format(elapsed_total_time) + "[sec]")
        

