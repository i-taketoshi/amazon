# -*- coding: utf-8 -*-
"""
Created on Sun Mar 11 02:07:41 2018

@author: taketoshi
"""

import urllib.parse
import urllib.request

import openpyxl as px

import time
import csv
from bs4 import BeautifulSoup

import sqlite3
import datetime
import re
#半角数字
digitReg = re.compile(r'^[0-9]+$')
def isdigit(s):
    return digitReg.match(str(s)) is not None

# アクセスするURL
url = "https://info.finance.yahoo.co.jp/fx/detail/?code=USDJPY=FX"

# URLにアクセスする htmlが帰ってくる → <html><head><title>経済、株価、ビジネス、政治のニュース:日経電子版</title></head><body....
html = urllib.request.urlopen(url)

# htmlをBeautifulSoupで扱う
soup = BeautifulSoup(html, "lxml")

strUSDJPY = soup.find(id='USDJPY_detail_bid')
USDJPY = float(strUSDJPY.text)



dbname = r'D:\workspace\sqlite\sample.sqlite3'
 #DB接続
conn = sqlite3.connect(dbname)
c = conn.cursor()


todaydetail  =    datetime.datetime.today()
yymmdd = todaydetail.strftime("%Y-%m-%d")
hhmmss = todaydetail.strftime("%H:%M:%S")
yymmdd = '2018-03-27'

#アマゾンテーブルから検索用のワードを検索する                            
sql = u"select jp_asin,JP_price,US_price from asins where yymmdd = '" + yymmdd + "'"
#sql = u"select jp_asin,JP_price,US_price from asins limit 10"
print(sql)
#sql = u"select jp_asin,JP_price,US_price from asins where US_price is not NULL limit 10"
c.execute(sql)
difflist = c.fetchall()

conn.commit()
conn.close()
Filename = r'D:\workspace\pickup_template - コピー.xlsm'
# ブックを開く
wb = px.load_workbook(Filename, read_only=False, keep_vba=True)
#シートを固定
ws = wb["diff"]
#最終行取得
max_row = ws.max_row
#項番最大値取得
maxno = ws.cell(row = ws.max_row, column = 1).value

if isdigit(maxno) is False:
    maxno = 0

#為替値代入             
ws['B1'].value = USDJPY

arr = []
rowcnt = 1

for i,d in enumerate(difflist):
    
    if d[2] is not None:
        us = d[2].replace('$','')

    try:
        #if (float(d[1])*0.9) > (float(us) * USDJPY):
        tmp = float(d[1])*0.9
        tmp = float(us) * USDJPY
        #if 1 > 0:
        
        #ASIN, JP, US
        arr.append([d[0],d[1],float(us) * USDJPY])
        
        maxno = maxno + 1
        incrow = max_row + rowcnt
        
        ws.cell(row = incrow,column = 1).value = maxno
        ws.cell(row = incrow,column = 2).value = d[0]
        ws.cell(row = incrow,column = 3).value = d[1]
        ws.cell(row = incrow,column = 4).value = us
        ws.cell(row = incrow,column = 5).value = "=$B$1*D" + str(incrow)
        
        ws.cell(row = incrow,column = 6).value = "=E" + str(incrow) + "/C" + str(incrow)
        ws.cell(row = incrow,column = 7).value = "=C" + str(incrow) + "-E" + str(incrow)
        rowcnt = rowcnt + 1
    except:
        #import traceback
        #traceback.print_exc()
        continue

    
wb.save(Filename)  




