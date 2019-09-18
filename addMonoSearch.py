# -*- coding: utf-8 -*-
"""
Created on Wed Apr 18 00:43:48 2018

@author: taketoshi

検索しきれなかったモノレート結果を検索する
xlsm対応
"""
#pip install xlrd

import pandas as pd
import openpyxl as px
from random import randint

from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
import time

import urllib.parse
import urllib.request

import openpyxl as px

import os
import sys

fname = "pickup2018-05-27.xlsm"
Filename = "D:\workspace\モノレート\\" + fname
# ブックを開く
wb = px.load_workbook(Filename, read_only=False, keep_vba=True)
#シートを固定
ws = wb["diff"]
#最終行取得
max_row = ws.max_row
#項番最大値取得
maxno = ws.cell(row = ws.max_row, column = 1).value


#為替値代入             
USDJPY = ws['B1'].value

arr = []
rowcnt = 1
monoratearr = []


sheetName = "diff"
df = pd.read_excel(Filename, skiprows=9, header=1)


df = df.fillna("#NA")

monoratearr = []

for i, v in df.iterrows():
    #print (i, v['ASIN'], v['JP価格'])    # v は Series
    #print(v['USドル'])
    if v['販売個数'] == "#NA":
        try:
            ratio = (v['USドル'] * USDJPY) / v['JP価格']
            #pricediff = v['JP価格'] - (v['USドル'] * USDJPY)
            #print(ratio)
            if ratio < 0.75 and ratio > 0.1:
                monoratearr.append([v['ASIN'],v['項番'] + 11])
                
        except ZeroDivisionError:
            #print("0だよ" + v['ASIN'])    
            pass



#print(monoratearr)

li_uniq = []
for x in monoratearr:
    if x[0] not in li_uniq:
        li_uniq.append(x)

print(li_uniq)

#########################
#モノレート
#########################


path = r'D:\workspace\amazon\chromedriver.exe'

options = Options()
options.add_argument("--disable-extentions")
options.add_argument("--ignore-certificate-errors")
#options.add_argument("--headless")


driver = webdriver.Chrome(path, chrome_options=options)
# ページの読み込み待ち時間(10秒)
#driver.set_page_load_timeout(10)

for u,asin in enumerate(li_uniq):
    
    #タイムアウト対策3回まで
    i = 0
    while i < 3:
        
        try:
            driver.get("https://www.mnrate.com/item/aid/" + asin[0])
            time.sleep(5)
        except TimeoutException:
            i = i + 1
            print("Timeout, Retrying... (%(i)s/%(max)s)" % {'i': i, 'max': 3})
            continue
        else:
            i = i + 1
            time.sleep(10)
            continue
    
    data = driver.page_source
    soup = BeautifulSoup(data,"lxml")
    
    #エラー処理
    if soup.find(class_='error_page_contents'):

        ws.cell(row = asin[1],column = 8).value = 'ASINなし'
        wb.save(Filename)
        print("情報なし")
        time.sleep(120)
        continue
    
    #エラー処理（データ取得不可時）
    try:
        
        item_price_sheet = soup.find(class_='item_price_sheet').get_text()
        if item_price_sheet.find('データを取得できませんでした') > 0:
            ws.cell(row = asin[1],column = 8).value = 'データ取得失敗'
            wb.save(Filename)
            print("取得失敗")
            time.sleep(60)
            continue
    except:
        ws.cell(row = asin[1],column = 8).value = 'None Type'
        wb.save(Filename)
        print("None Type")
        time.sleep(30)
        continue


    #テーブルを指定
    table = soup.find(id="sheet_contents")
    rows = table.findAll("tr")
    
    rank = []
    try:
        for row in rows:
            csvRow = []
            for cell in row.findAll(['td', 'th']):
                csvRow.append(str.strip(cell.get_text()))
            
            if len(csvRow) > 1 and  csvRow[1] != "":
                    rank.append(int(csvRow[1]))
    #        writer.writerow(csvRow)
    except:
        pass
        #csvFile.close()
    
    
    #販売個数取得
    cnt = 0
    cnt2 = 0
    
    if len(rank) != 0:
        
        mn = min(rank)
        
        ranklist = ','.join(map(str, rank))
        #谷の数を数える
        for i in range(len(rank)-2):
            #1回売れた場合
            if (rank[i+1] < rank[i]) and (rank[i+1] < rank[i+2]):
                cnt += 1
            #２連続で売れた場合
            if rank[i+1] > rank[i+2]:
                cnt += 1
            
            #前日よりランキングが2割下がった
            if rank[i]*0.8 > rank[i+1]:
                cnt2+=1
        
        #for i in range(len(rank)-1):
        #   #前日よりランキングが上(数値指摘には下)ならカウント
        #    if rank[i] > rank[i+1]:
        #        cnt+=1
        #    
        #    #前日よりランキングが2割下がった
        #    if rank[i]*0.8 > rank[i+1]:
        #        cnt2+=1
        
        #print("---------")
        #print("rank上昇回数：" + str(cnt))
        #print("rank下落率が平均値の1割以内：" + str(cnt2))
    
        ws.cell(row = asin[1],column = 8).value = str(cnt)
        ws.cell(row = asin[1],column = 9).value = str(cnt2)
        ws.cell(row = asin[1],column = 10).value = str(mn)
        ws.cell(row = asin[1],column = 11).value = str(ranklist)
        
    else:
        ws.cell(row = asin[1],column = 8).value = 0
        
    print(str(u+1) + "/" + str(len(monoratearr))+ "個目")   
    wb.save(Filename)
    
    
    

    time.sleep(randint(10,30))
    hp = randint(1,3)
    i = 0
    while i < 3:
        
        try:
            if hp == 1:
                driver.get("https://www.yahoo.co.jp/")
            elif hp == 2:
                driver.get("https://www.google.co.jp/")
            #elif hp == 3:
            #    driver.get("https://www.excite.co.jp/")
            else:
                driver.get("https://www.bing.com/")
    
            time.sleep(randint(6,14))
        except TimeoutException:
            i = i + 1
            print("Timeout, Retrying... (%(i)s/%(max)s)" % {'i': i, 'max': 3})
            continue
        else:
            i = i + 1
            time.sleep(5)
            continue
                           
driver.quit()